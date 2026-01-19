import re
import datetime as dt
from pathlib import Path

import pandas as pd
import streamlit as st
import openpyxl


APP_DIR = Path(__file__).parent
XLSM_PATH = APP_DIR / "schade met macro.xlsm"
LOGO_PATH = APP_DIR / "logo.png"
SHEET_NAME = "BRON"

REQUIRED_COLS = [
    "personeelsnr",
    "volledige naam",
    "Datum",
    "Link",
    "Locatie",
    "voertuig",
    "bus/tram",
    "type",
]


def norm(s) -> str:
    return str(s).strip().lower()


def parse_year(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, (dt.date, dt.datetime)):
        return v.year

    s = str(v).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        return int(m.group(3))

    m2 = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m2:
        return int(m2.group(1))

    try:
        return dt.datetime.fromisoformat(s).year
    except Exception:
        return None


@st.cache_data(show_spinner=False)
def load_bron_df() -> pd.DataFrame:
    if not XLSM_PATH.exists():
        raise FileNotFoundError(f"Bestand niet gevonden: {XLSM_PATH.name}")

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Tabblad '{SHEET_NAME}' niet gevonden in {XLSM_PATH.name}")

    ws = wb[SHEET_NAME]

    header = [c.value for c in ws[1]]
    header_map = {norm(h): idx for idx, h in enumerate(header)}

    def find_idx(col: str) -> int | None:
        key = norm(col)
        if key in header_map:
            return header_map[key]
        # tolerant voor varianten
        if col == "bus/tram":
            for alt in ["bus/ tram", "bus / tram", "bus - tram"]:
                if alt in header_map:
                    return header_map[alt]
        if col == "volledige naam":
            for alt in ["naam", "volledige naam.", "volledige naam "]:
                if alt in header_map:
                    return header_map[alt]
        return None

    idx_map = {c: find_idx(c) for c in REQUIRED_COLS}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        obj = {}
        any_val = False
        for col in REQUIRED_COLS:
            i = idx_map.get(col)
            val = r[i] if (i is not None and i < len(r)) else None

            if val is not None and str(val).strip() != "":
                any_val = True

            if col == "Datum" and isinstance(val, (dt.date, dt.datetime)):
                val = val.isoformat()

            obj[col] = val

        if any_val:
            rows.append(obj)

    df = pd.DataFrame(rows)

    # Zorg dat alle vereiste kolommen bestaan (ook als leeg)
    for c in REQUIRED_COLS:
        if c not in df.columns:
            df[c] = None

    df["_jaar"] = df["Datum"].apply(parse_year)

    # zoekveld index
    df["_search"] = (
        df["personeelsnr"].fillna("").astype(str) + " " +
        df["volledige naam"].fillna("").astype(str) + " " +
        df["voertuig"].fillna("").astype(str)
    ).str.lower()

    return df


def build_suggestions(df: pd.DataFrame, q: str, limit: int = 10) -> list[str]:
    q = (q or "").strip().lower()
    if not q:
        return []

    candidates = []

    # unieke suggesties uit 3 velden
    for col in ["personeelsnr", "volledige naam", "voertuig"]:
        vals = df[col].dropna().astype(str).unique().tolist()
        for v in vals:
            if q in v.lower():
                candidates.append(v)

    # dedup + limit
    seen = set()
    out = []
    for v in candidates:
        key = v.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(v)
        if len(out) >= limit:
            break
    return out


# =========================
# Streamlit UI
# =========================
st.set_page_config(page_title="Analyse en rapportering OT Gent", layout="wide")

# Sidebar
with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), use_container_width=True)
    st.markdown("### Analyse en rapportering OT Gent")
    st.caption("schade")

# Load data
try:
    df = load_bron_df()
except Exception as e:
    st.error(f"Kan data niet laden: {e}")
    st.stop()

# Jaarfilter
years = sorted([y for y in df["_jaar"].dropna().unique().tolist() if y is not None], reverse=True)
with st.sidebar:
    year_choice = st.selectbox("Jaar", ["Alle"] + [str(y) for y in years], index=0)

if year_choice != "Alle":
    df_view = df[df["_jaar"] == int(year_choice)].copy()
else:
    df_view = df.copy()

# Top menu (tabs)
tab_dashboard, tab_chauffeur, tab_voertuig, tab_locatie, tab_coaching, tab_analyse = st.tabs(
    ["Dashboard", "Chauffeur", "Voertuig", "Locatie", "Coaching", "Analyse"]
)

# Dashboard: zoek + suggesties
with tab_dashboard:
    st.subheader("Dashboard")

    if "q" not in st.session_state:
        st.session_state.q = ""

    q = st.text_input(
        "Zoek op personeelsnr, volledige naam of voertuig",
        value=st.session_state.q,
        placeholder="Typ om te zoeken…",
        key="q_input",
    )

    suggestions = build_suggestions(df_view, q, limit=10)

    # Dropdown met suggesties (bij typen)
    sel = st.selectbox(
        "Suggesties",
        options=[""] + suggestions,
        index=0,
        help="Klik een suggestie om je zoekveld te vullen.",
    )

    if sel:
        st.session_state.q = sel
        st.rerun()
    else:
        st.session_state.q = q

    q_norm = (st.session_state.q or "").strip().lower()

    if q_norm:
        hits = df_view[df_view["_search"].str.contains(re.escape(q_norm), na=False)].copy()
    else:
        hits = df_view.copy()

    st.caption(f"Records: {len(hits)} (jaarfilter: {year_choice})")

    # Toon alleen jouw kolommen
    hits_show = hits[REQUIRED_COLS].head(200).copy()

    st.data_editor(
        hits_show,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Link": st.column_config.LinkColumn("Link"),
        },
        disabled=True,
    )

# Placeholders voor andere tabs (kan je later vullen)
with tab_chauffeur:
    st.info("Chauffeur: later uitwerken (filters/aggregaties op BRON).")

with tab_voertuig:
    st.info("Voertuig: later uitwerken (top voertuigen, trends, …).")

with tab_locatie:
    st.info("Locatie: later uitwerken (top locaties, heatmap, …).")

with tab_coaching:
    st.info("Coaching: later uitwerken (koppeling met Coachingslijst.xlsx en gesprekken).")

with tab_analyse:
    st.info("Analyse: later uitwerken (grafieken per maand, schade per type, …).")
